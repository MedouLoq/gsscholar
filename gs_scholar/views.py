from django.shortcuts import render
from django.urls import reverse
from django.http import HttpResponseBadRequest
import zipfile
from django.contrib.auth.decorators import login_required
from matplotlib.ticker import MaxNLocator
# Create your views here.
from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404, redirect
from .models import *
from openpyxl import Workbook
from .forms import *
from .forms import SectionForm, QuestionForm ,EmailtableForm
from django.utils import timezone
from datetime import timedelta ,date
from matplotlib.ticker import MaxNLocator
import seaborn as sns
from dateutil import parser
import numpy as np
from django.conf import settings
import pandas as pd
import matplotlib.pyplot as plt
import io
import openpyxl
import os
from math import pi
from django.views import View
from django.http import HttpResponse
from django.core.mail import send_mail
import base64
from django.core.mail import EmailMessage
from django.template.loader import render_to_string

def is_admin_required(function):

    def wrapper(request, *args, **kwargs):
        if not request.user.is_superuser:
            return redirect('student_home')
        return function(request, *args, **kwargs)
    return wrapper

@login_required
def student_home(request):
    # Superusers are redirected to admin dashboard
    if request.user.is_superuser:
        return redirect('index')
    try:
        student = Etudiant.objects.get(user=request.user)
    except Etudiant.DoesNotExist:
        # If no student record exists, redirect to a safe error page or show an error.
        return redirect('error_page')
    # Retrieve pending and completed questionnaire info.
    studentinfo_pending = Studentinfo.objects.filter(student=student, completed=False)
    studentinfo_completed = Studentinfo.objects.filter(student=student, completed=True)
    context = {
        'student': student,
        'studentinfo': studentinfo_pending,
        'studentinfo2': studentinfo_completed,
    }
    return render(request, 'student_profile.html', context)


def custom_404(request, exception):
    return render(request, 'technical_404.html', status=404)

def custom_500(request):
    return render(request, 'technical_500.html', status=500)


# Create your views here.
@is_admin_required
@login_required
def Prof_table(request):
  promo = Proffesseur.objects.all()
  context = {
      'promo': promo
  }
  return render(request, 'Module_table.html', context)
@is_admin_required
@login_required
def add_module(request):
  if request.method == 'POST':
    module=Proffesseur(
    id = request.POST['id'],
    Nom_du_Prof = request.POST['Nom_du_Prof'],
    numero_tel =request.POST['numero_tel'],
    email =request.POST['email'],
    Specialite =request.POST['specialite'],)

    module.save()
    return redirect('Module_table')
  context = {}
  return render(request, 'add_module.html', context)
@is_admin_required
@login_required
def update_module(request, id):
  module = get_object_or_404(Proffesseur, id=id)
  if request.method == 'POST':
    module.id = request.POST['id']
    module.numero = request.POST['numero']
    module.module_name = request.POST['module_name']
    module.save()
    return redirect('Module_table')
  context = {'module': module}
  return render(request, 'update_module.html', context)
@is_admin_required
@login_required
def delete_module(request, id):
  module = get_object_or_404(Proffesseur, id=id)
  module.delete()
  return redirect('Module_table')



# -----------------------
# MATIER VIEWS
# -----------------------

@is_admin_required
@login_required
def matier_table(request):
    # Use select_related to load related Semestre in one query
    matiers = Matier.objects.select_related('semester').all()
    context = {
        'matiers': matiers,
    }
    return render(request, 'matier_table.html', context)


@login_required
def add_matier(request):
    if request.method == 'POST':
        try:
            semester = Semestre.objects.get(pk=request.POST['semester'])
        except Semestre.DoesNotExist:
            return render(request, 'add_matier.html', {
                'errors': "Selected semester does not exist.",
                'specialites': Specialite.objects.all(),
                'semestres': Semestre.objects.all()
            })

        specialite_id = request.POST.get('specialite')
        specialite = Specialite.objects.get(pk=specialite_id) if specialite_id else None
        is_common = request.POST.get('is_common') == 'on'

        matier = Matier(
            id=request.POST['id'],
            name=request.POST['name'],
            code=request.POST['code'],
            semester=semester,
            is_common=is_common,
            specialite=specialite
        )
        try:
            matier.full_clean()
            matier.save()
        except ValidationError as ve:
            return render(request, 'add_matier.html', {
                'errors': ve.message_dict,
                'specialites': Specialite.objects.all(),
                'semestres': Semestre.objects.all()
            })
        return redirect('matier_table')
    else:
        context = {
            'specialites': Specialite.objects.all(),
            'semestres': Semestre.objects.all(),
        }
        return render(request, 'add_matier.html', context)


@login_required
def update_matier(request, id):
    matier = get_object_or_404(Matier, pk=id)
    if request.method == 'POST':
        matier.name = request.POST['name']
        matier.code = request.POST['code']
        matier.is_common = request.POST.get('is_common') == 'on'
        try:
            matier.semester = Semestre.objects.get(pk=request.POST['semester'])
        except Semestre.DoesNotExist:
            return render(request, 'update_matier.html', {
                'errors': "Selected semester does not exist.",
                'matier': matier,
                'specialites': Specialite.objects.all(),
                'semestres': Semestre.objects.all()
            })
        specialite_id = request.POST.get('specialite')
        matier.specialite = Specialite.objects.get(pk=specialite_id) if specialite_id else None

        try:
            matier.full_clean()
            matier.save()
        except ValidationError as ve:
            return render(request, 'update_matier.html', {
                'errors': ve.message_dict,
                'matier': matier,
                'specialites': Specialite.objects.all(),
                'semestres': Semestre.objects.all()
            })
        return redirect('matier_table')
    else:
        context = {
            'matier': matier,
            'specialites': Specialite.objects.all(),
            'semestres': Semestre.objects.all(),
        }
        return render(request, 'update_matier.html', context)


@login_required
def delete_matier(request, id):
    matier = get_object_or_404(Matier, pk=id)
    matier.delete()
    return redirect('matier_table')


# -----------------------
# ETUDIANT VIEWS
# -----------------------

@login_required
@is_admin_required  # Use your decorator
def etudiant_table(request):
    """
    Displays and filters the Etudiant table.  Handles L1, L2, and L3 separately,
    with specialization filtering for L2 and L3.
    """
    niveau_filter = request.GET.get('niveau')
    specialite_filter = request.GET.get('specialite')

    if niveau_filter:  # If a specific level is selected
        etudiants = Etudiant.objects.filter(niveau__nom_niveau=niveau_filter)

        # Filter by specialite ONLY if a specialite is selected AND we are on L2/L3
        if specialite_filter and niveau_filter in ('L2', 'L3'):
            etudiants = etudiants.filter(specialite__id=specialite_filter)

        # Get specialties for the filter dropdown (only for L2/L3)
        specialties = []
        if niveau_filter in ('L2', 'L3'):
          niveaux = Niveau.objects.filter(nom_niveau=niveau_filter)
          specialties = Specialite.objects.filter(niveau__in=niveaux) # Get specialties related to the Niveau


        context = {
            f'etudiants_{niveau_filter}': etudiants,  # Use dynamic context variable name
            'specialties': specialties,  # Always pass specialties, even if empty
            'niveau_filter': niveau_filter, # Pass the filter for the template use.
        }
        return render(request, 'etudiant_table.html', context)

    else:  # No specific level selected, show all levels
        etudiants_L1 = Etudiant.objects.filter(niveau__nom_niveau='L1')
        etudiants_L2 = Etudiant.objects.filter(niveau__nom_niveau='L2')
        etudiants_L3 = Etudiant.objects.filter(niveau__nom_niveau='L3')
        # Pass specialties for L2 and L3 filters, even on initial load.
        specialties_L2 = Specialite.objects.filter(niveau__nom_niveau='L2')
        specialties_L3 = Specialite.objects.filter(niveau__nom_niveau='L3')
        context = {
            'etudiants_L1': etudiants_L1,
            'etudiants_L2': etudiants_L2,
            'etudiants_L3': etudiants_L3,
            'specialties_L2': specialties_L2,  # Specialties for L2
            'specialties_L3': specialties_L3,  # Specialties for L3
        }
        return render(request, 'etudiant_table.html', context)


@is_admin_required
@login_required
def add_etudiant(request):
    if request.method == 'POST':
        try:
            niveau = Niveau.objects.get(pk=request.POST['niveau'])
        except Niveau.DoesNotExist:
            return render(request, 'add_etudiant.html', {
                'errors': "Selected niveau does not exist.",
                'niveaux': Niveau.objects.all(),
                'specialites': Specialite.objects.all()
            })
        specialite_id = request.POST.get('specialite')
        specialite = Specialite.objects.get(pk=specialite_id) if specialite_id else None

        etudiant = Etudiant(
            Matricul=request.POST['Matricul'],
            nom_etudiant=request.POST['nom_etudiant'],
            email=request.POST['email'],
            sexe=request.POST['sexe'],
            niveau=niveau,
            specialite=specialite,
        )
        try:
            etudiant.full_clean()
            etudiant.save()
        except ValidationError as ve:
            return render(request, 'add_etudiant.html', {
                'errors': ve.message_dict,
                'niveaux': Niveau.objects.all(),
                'specialites': Specialite.objects.all()
            })
        return redirect('etudiant_table')
    else:
        context = {
            'niveaux': Niveau.objects.all(),
            'specialites': Specialite.objects.all(),
        }
        return render(request, 'add_etudiant.html', context)


@is_admin_required
@login_required
def update_etudiant(request, id):
    etudiant = get_object_or_404(Etudiant, pk=id)
    if request.method == 'POST':
        etudiant.nom_etudiant = request.POST['nom_etudiant']
        etudiant.email = request.POST['email']
        etudiant.sexe = request.POST['sexe']
        etudiant.Matricul = request.POST['Matricul']
        try:
            etudiant.niveau = Niveau.objects.get(pk=request.POST['niveau'])
        except Niveau.DoesNotExist:
            return render(request, 'update_etudiant.html', {
                'errors': "Selected niveau does not exist.",
                'etudiant': etudiant,
                'niveaux': Niveau.objects.all(),
                'specialites': Specialite.objects.all()
            })
        specialite_id = request.POST.get('specialite')
        etudiant.specialite = Specialite.objects.get(pk=specialite_id) if specialite_id else None

        try:
            etudiant.full_clean()
            etudiant.save()
        except ValidationError as ve:
            return render(request, 'update_etudiant.html', {
                'errors': ve.message_dict,
                'etudiant': etudiant,
                'niveaux': Niveau.objects.all(),
                'specialites': Specialite.objects.all()
            })
        return redirect('etudiant_table')
    else:
        context = {
            'etudiant': etudiant,
            'niveaux': Niveau.objects.all(),
            'specialites': Specialite.objects.all(),
        }
        return render(request, 'update_etudiant.html', context)


@is_admin_required
@login_required
def delete_etudiant(request, Matricul):
    student = get_object_or_404(Etudiant, pk=Matricul)
    student.delete()
    return redirect('etudiant_table')


@is_admin_required
@login_required
def import_etudiants(request):
    # If a 'niveau' GET parameter is passed, only import for that level.
    niveau_code = request.GET.get('niveau')  # expecting values like 'L1', 'L2', or 'L3'
    if request.method == 'POST':
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                df = pd.read_excel(excel_file)
                for index, row in df.iterrows():
                    # If niveau_code is provided, override the Excel value
                    if niveau_code:
                        nom_niveau = niveau_code
                    else:
                        nom_niveau = row['Niveau']
                    niveau_instance = Niveau.objects.get(nom_niveau=nom_niveau)
                    
                    # For L2 and L3, try to get the specialité from the Excel row.
                    specialite_instance = None
                    if nom_niveau in ['L2', 'L3']:
                        # Adjust the column name as per your Excel file – here we assume "Specialite"
                        specialite_value = row.get('Specialite', None)
                        if specialite_value:
                            try:
                                # You might use code or name to lookup the specialité.
                                specialite_instance = Specialite.objects.get(code=specialite_value)
                            except Specialite.DoesNotExist:
                                # Optionally, try to get it by name or log an error.
                                try:
                                    specialite_instance = Specialite.objects.get(name__iexact=specialite_value)
                                except Specialite.DoesNotExist:
                                    specialite_instance = None

                    etudiant = Etudiant(
                        nom_etudiant=row['Prénom & Nom'],
                        Matricul=row['Matricule'],
                        sexe=row['Sexe'],
                        email=row['E-mail'],
                        niveau=niveau_instance,
                        specialite=specialite_instance
                    )
                    etudiant.save()
                return redirect('etudiant_table')
            except Exception as e:
                form.add_error(None, 'Error processing the Excel file. Please check the format.')
    else:
        form = StudentImportForm()

    return render(request, 'import_etudiants.html', {'form': form, 'niveau': niveau_code})

#AnneeScolaire
@is_admin_required
@login_required
def annee_scolaire_table(request):
  annee_scolaires = AnneeScolaire.objects.all().order_by('-annee')
  context = {
    'annee_scolaires': annee_scolaires,
  }
  return render(request, 'annee_scolaire_table.html', context)


@is_admin_required
@login_required
def add_annee_scolaire(request):
  if request.method == 'POST':
    annee_scolaire = AnneeScolaire(
        annee=request.POST['annee'],
        current=request.POST.get('current') == 'on'
    )
    annee_scolaire.save()
    return redirect('annee_scolaire_table')
  context = {}
  return render(request, 'add_annee_scolaire.html', context)


@is_admin_required
@login_required
def update_annee_scolaire(request, id):
    annee = get_object_or_404(AnneeScolaire, pk=id)

    if request.method == 'POST':
        annee.current = request.POST.get('current') == 'on'
        annee.save()
        return redirect('annee_scolaire_table')

    return render(request, 'update_annee_scolaire.html', {'annee': annee})

@is_admin_required
@login_required
def delete_annee_scolaire(request, annee_scolaire):
  student = get_object_or_404(AnneeScolaire, pk=annee_scolaire)
  student.delete()
  return redirect('annee_scolaire_table')

#Niveau


@is_admin_required
@login_required
def niveau_table(request):
  niveau = Niveau.objects.all()
  context = {
    'niveau': niveau,
  }
  return render(request, 'niveau_table.html', context)


@is_admin_required
@login_required
def add_niveau(request):
  if request.method == 'POST':
    niveau = Niveau(
        niveau=request.POST['niveau'],
    )
    niveau.save()
    return redirect('niveau_table')
  context = {}
  return render(request, 'add_niveau.html', context)

@is_admin_required
@login_required
def update_niveau(request, niveau):
  niveau_list = Proffesseur.objects.all()  # This line is problematic
  niveau = get_object_or_404(Niveau, pk=niveau)  # Corrected to Niveau
  if request.method == 'POST':
    niveau.niveau = request.POST['niveau']
    niveau.save()
    return redirect('niveau_table')
  context = {'niveau_list': niveau_list, 'niveau': niveau}
  return render(request, 'update_niveau.html', context)

@is_admin_required
@login_required
def delete_niveau(request, niveau):
  student = get_object_or_404(Niveau, pk=niveau)
  student.delete()
  return redirect('niveau_table')

#Semestre
@is_admin_required
@login_required
def semestre_table(request):
  semestre = Semestre.objects.all()
  context = {
   'semestre': semestre,
  }
  return render(request,'semestre_table.html', context)

@is_admin_required
@login_required
def add_semestre(request):
  if request.method == 'POST':
    semestre = Semestre(
        semestre=request.POST['semestre'],
    )
    semestre.save()
    return redirect('semestre_table')
  context = {}
  return render(request, 'add_semestre.html', context)


@is_admin_required
@login_required
def update_semestre(request, semestre):
  semestre_list = Proffesseur.objects.all()  # This line is problematic
  semestre = get_object_or_404(Semestre, pk=semestre)  # Corrected to Semestre
  if request.method == 'POST':
    semestre.semestre = request.POST['semestre']
    semestre.save()
    return redirect('semestre_table')
  context = {'semestre_list': semestre_list,'semestre': semestre}
  return render(request, 'update_semestre.html', context)


@is_admin_required
@login_required
def delete_semestre(request, semestre):
  student = get_object_or_404(Semestre, pk=semestre)
  student.delete()
  return redirect('semestre_table')

@is_admin_required
@login_required
def index(request):
    anne=AnneeScolaire.objects.get(current=True)
    return render(request, 'index2.html',{'anne':anne})

@is_admin_required
@login_required
def questionnaire(request,questionnaire_id):
   questionnaire = Questionnaire.objects.get(pk=questionnaire_id)
   sections = questionnaire.sections.all()
   matiere = questionnaire.matier
   return render(request,'questionaire.html',{'questionnaire': questionnaire, 'sections': sections,'matiere':matiere})



import base64
from asgiref.sync import sync_to_async
from django.http import HttpResponseBadRequest, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views import View
from datetime import timedelta
from gs_scholar.models import Questionnaire, Etudiant, AnneeScolaire, Studentinfo, Datainserted, DataInsertionManager

class QuestionnaireView(View):
    template_name = 'questionaire.html'

    async def get(self, request, questionnaire_id):
        questionnaire = await sync_to_async(get_object_or_404)(Questionnaire, pk=questionnaire_id)
        student_identifier_encoded = request.GET.get('student_identifier')
        if not student_identifier_encoded:
            return HttpResponseBadRequest("Student identifier is missing in the URL.")

        expiration_time = questionnaire.expiration_time
        if expiration_time and timezone.now() > expiration_time:
            return await sync_to_async(render)(request, 'temp_expired.html')

        try:
            student_identifier = base64.b64decode(student_identifier_encoded).decode('utf-8')
            student = await sync_to_async(get_object_or_404)(Etudiant, email=student_identifier)
        except Exception:
            return HttpResponseBadRequest("Invalid student identifier.")

        # Use sync_to_async to access lazy fields:
        student_niveau = await sync_to_async(lambda: student.niveau)()
        questionnaire_semester_niveau = await sync_to_async(lambda: questionnaire.semester.niveau)()
        if student_niveau != questionnaire_semester_niveau:
            return HttpResponse("Access denied: Your level does not match this questionnaire.", status=403)
        if student.niveau.nom_niveau in ['L2', 'L3']:
            # Access speciality safely
            student_speciality = await sync_to_async(lambda: student.specialite)()
            questionnaire_speciality = await sync_to_async(lambda: questionnaire.matier.specialite)()
            if not questionnaire.matier.is_common and student_speciality != questionnaire_speciality:
                return HttpResponse("Access denied: Your specialty does not match this questionnaire.", status=403)

        if await sync_to_async(Datainserted.objects.filter(questionnaire=questionnaire, etudiant=student).exists)():
            return redirect('stop')

        return await sync_to_async(render)(request, self.template_name, {'questionnaire': questionnaire, 'student': student})

    async def post(self, request, questionnaire_id):
        questionnaire = await sync_to_async(get_object_or_404)(Questionnaire, pk=questionnaire_id)
        sections = await sync_to_async(list)(questionnaire.sections.all())
        student_identifier_encoded = request.GET.get('student_identifier')
        if not student_identifier_encoded:
            return HttpResponseBadRequest("Student identifier is missing in the URL.")
    
        try:
            student_identifier = base64.b64decode(student_identifier_encoded).decode('utf-8')
            student = await sync_to_async(get_object_or_404)(Etudiant, email=student_identifier)
        except Exception:
            return HttpResponseBadRequest("Invalid student identifier.")
    
        total_questions = 0
        responses_data = []
        anne = await sync_to_async(get_object_or_404)(AnneeScolaire, current=True)
        
        # Wrap access to questionnaire.matier:
        matier_obj = await sync_to_async(lambda: questionnaire.matier)()
        matier_code = matier_obj.code if matier_obj else 'Default'
        
        for section in sections:
            questions = await sync_to_async(list)(section.questions.all())
            total_questions += len(questions)
            for question in questions:
                field_name = f"question_{question.id}"
                selected_choice = request.POST.get(field_name)
                if not selected_choice:
                    continue
                weight = question.get_weight_for_choice(selected_choice)
                responses_data.append({
                    'questionnaire_id': questionnaire.id,
                    'question_id': question.id,
                    'section_id': section.id,
                    'matier_code': matier_code,
                    'weight': weight,
                    'selected_choice': selected_choice,
                    'etudiant_id': student.Matricul,
                    'anne_id': anne.id,
                    'specialite_id': student.specialite_id if student.specialite else None
                })
        
        # Process all responses in bulk
        await DataInsertionManager.process_questionnaire_responses(responses_data, Datainserted)
        
        # If all questions were answered, update student info
        if len(responses_data) == total_questions:
            await sync_to_async(Studentinfo.objects.update_or_create)(
                student=student,
                matier=questionnaire.name,
                defaults={'completed': True}
            )
        
        return redirect('stop')

def stop(request):
    return render(request,'stop.html')


from datetime import timedelta
import base64
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.views import View
from django.db.models import Q, F, ExpressionWrapper, IntegerField, Max
from django.db.models.functions import Mod
from django.contrib.auth.decorators import login_required
from gs_scholar.models import (
    Questionnaire, Semestre, Specialite, Niveau, Etudiant, Studentinfo
)


# We'll use a custom dispatch form – here, for simplicity, we’ll process the POST data manually.
@is_admin_required
@login_required
def mail_sender(request):
    """
    Display a dispatch form that lets you select:
     - Multiple Semesters
     - Multiple Specialties (optional)
     - Whether to include common courses
     - Custom expiration time (in hours)
    """
    # Gather data for the form:
    semesters = Semestre.objects.all().order_by('nom_semestre')
    specialties = Specialite.objects.all().order_by('name')
    context = {
        'semesters': semesters,
        'specialties': specialties,
    }
    return render(request, 'mail_sender.html', context)


@login_required
def mail_sender_json(request):
    """
    Existing JSON view to fetch expiration times.
    (You may keep this if needed for a live countdown.)
    """
    now = timezone.now()
    questionnaires = Questionnaire.objects.annotate(
        mod_order=ExpressionWrapper(Mod(F('semester__order'), 2), output_field=IntegerField())
    )
    latest_impaired = questionnaires.filter(mod_order=1, expiration_time__gte=now).aggregate(max_expiration=Max('expiration_time'))
    latest_paired = questionnaires.filter(mod_order=0, expiration_time__gte=now).aggregate(max_expiration=Max('expiration_time'))
    impaired_time = latest_impaired.get('max_expiration')
    paired_time = latest_paired.get('max_expiration')
    chosen_time = None
    semester_type = None
    if impaired_time and impaired_time >= now:
        chosen_time = impaired_time
        semester_type = 'impaired'
    if paired_time and paired_time >= now:
        if chosen_time is None or paired_time > chosen_time:
            chosen_time = paired_time
            semester_type = 'paired'
    if chosen_time is None:
        return JsonResponse({'error': 'No active questionnaire found'}, status=404)
    return JsonResponse({'expiration_time': chosen_time.strftime('%Y-%m-%d %H:%M:%S'),
                         'semester_type': semester_type})


class SendQuestionnaireEmailView(View):
    def post(self, request):
        # Retrieve posted data:
        selected_semester_ids = request.POST.getlist('semesters')
        selected_specialties = request.POST.getlist('specialties')  # Optional
        include_common = request.POST.get('include_common') == 'on'
        try:
            expiration_hours = float(request.POST.get('expiration_time'))
        except (TypeError, ValueError):
            return JsonResponse({'error': 'Invalid expiration time'}, status=400)
        expiration_time = timezone.now() + timedelta(hours=expiration_hours)

        # Build a queryset for questionnaires in the selected semesters:
        qs = Questionnaire.objects.filter(semester__id__in=selected_semester_ids)
        if selected_specialties:
            qs = qs.filter(matier__specialite__id__in=selected_specialties)
        if include_common:
            common_qs = Questionnaire.objects.filter(
                semester__id__in=selected_semester_ids,
                matier__is_common=True
            )
            qs = (qs | common_qs).distinct()

        # Update expiration time for each questionnaire:
        for questionnaire in qs:
            questionnaire.expiration_time = expiration_time
            questionnaire.save()

        # Process each selected semester separately:
        selected_semesters = Semestre.objects.filter(id__in=selected_semester_ids)
        for semester in selected_semesters:
            # Select students from the semester's niveau:
            students = Etudiant.objects.filter(niveau=semester.niveau)

            # For L2/L3, ensure students have the correct speciality.
            if semester.niveau.nom_niveau in ['L2', 'L3']:
                students = students.filter(specialite=semester.niveau.specialite)

            # Get questionnaires specifically for this semester:
            semester_questionnaires = qs.filter(semester=semester)
            for student in students:
                for questionnaire in semester_questionnaires:
                    # For L2/L3, skip if not common and student’s speciality doesn’t match.
                    if semester.niveau.nom_niveau in ['L2', 'L3']:
                        if not questionnaire.matier.is_common and questionnaire.matier.specialite != student.specialite:
                            continue

                    student_identifier_encoded = base64.b64encode(student.email.encode('utf-8')).decode('utf-8')
                    questionnaire_url = request.build_absolute_uri(
                        reverse('display_questionnaire', args=[questionnaire.id])
                    ) + f'?student_identifier={student_identifier_encoded}'

                    # Store the URL in Studentinfo (tracking the speciality, too)
                    studentinfo, created = Studentinfo.objects.get_or_create(
                        student=student,
                        matier=questionnaire.name,
                        specialite=student.specialite
                    )
                    studentinfo.url = questionnaire_url
                    studentinfo.save()

                    # Send the email:
                    subject = "Notification d'envoi de formulaire"
                    from_email = 'your_email@example.com'
                    to_email = [student.email]
                    html_content = render_to_string('mail2.html', {
                        'first_name': student.nom_etudiant,
                        'last_name': student.Matricul,
                        'platform_link': 'your_platform_link',
                        'username': student.email,
                        'password': '************',
                        'semestre': semester,
                    })
                    email = EmailMessage(subject, html_content, from_email, to_email)
                    email.content_subtype = 'html'
                    email.send()
        return JsonResponse({'status': 'success'})


@is_admin_required
@login_required
def questionnaire_table(request):
    questionnaires = Questionnaire.objects.all()
    context = {'questionnaires': questionnaires}
    return render(request, 'questionnaire_table.html', context)

@is_admin_required
@login_required
def add_questionnaire(request):
    if request.method == 'POST':
        name = request.POST['name']
        description = request.POST['description']
        professor_id = request.POST['professor']
        matier_id = request.POST['matier']
        semester_id = request.POST['semester']

        # Retrieve selected sections from checkboxes
        sections_ids = request.POST.getlist('sections')

        professor = get_object_or_404(Proffesseur, pk=professor_id)
        matier = get_object_or_404(Matier, pk=matier_id)
        semester = get_object_or_404(Semestre, pk=semester_id)

        questionnaire = Questionnaire(
            name=name,
            description=description,
            professor=professor,
            matier=matier,
            semester=semester,
        )
        questionnaire.save()

        # Add selected sections to the questionnaire
        questionnaire.sections.set(sections_ids)

        return redirect('questionnaire_table')

    professors = Proffesseur.objects.all()
    matiers = Matier.objects.all()
    semesters = Semestre.objects.all()
    sections = Section.objects.all()

    context = {
        'professors': professors,
        'matiers': matiers,
        'semesters': semesters,
        'sections': sections,
    }
    return render(request, 'add_questionnaire.html', context)

@is_admin_required
@login_required
def get_matiers_for_semester(request, semester_id):
    # Assuming Matier model has a ForeignKey field named 'semester'
    matiers = Matier.objects.filter(semester_id=semester_id).values('id', 'name')
    return JsonResponse(list(matiers), safe=False)

@is_admin_required
@login_required
def update_questionnaire(request, id):
    questionnaire = get_object_or_404(Questionnaire, pk=id)

    if request.method == 'POST':
        questionnaire.name = request.POST['name']
        questionnaire.description = request.POST['description']

        professor_id = request.POST['professor']
        matier_id = request.POST['matier']
        semester_id = request.POST['semester']

        # Retrieve selected sections from checkboxes
        sections_ids = request.POST.getlist('sections')

        professor = get_object_or_404(Proffesseur, pk=professor_id)
        matier = get_object_or_404(Matier, pk=matier_id)
        semester = get_object_or_404(Semestre, pk=semester_id)

        questionnaire.professor = professor
        questionnaire.matier = matier
        questionnaire.semester = semester

        # Update selected sections for the questionnaire
        questionnaire.sections.set(sections_ids)

        questionnaire.save()
        return redirect('questionnaire_table')

    professors = Proffesseur.objects.all()
    matiers = Matier.objects.all()
    semesters = Semestre.objects.all()
    sections = Section.objects.all()

    context = {
        'professors': professors,
        'matiers': matiers,
        'semesters': semesters,
        'sections': sections,
        'questionnaire': questionnaire,
    }
    return render(request, 'update_questionnaire.html', context)

@is_admin_required
@login_required
def delete_questionnaire(request, id):
    questionnaire_instance = get_object_or_404(Questionnaire, pk=id)
    questionnaire_instance.delete()
    return redirect('questionnaire_table')

@is_admin_required
@login_required
def questionnaire_list(request,id):
    semester = Semestre.objects.all().get(id=id)
    questionnaires = Questionnaire.objects.all().filter(semester=semester)
    context = {'questionnaires': questionnaires}
    return render(request, 'questionnaire_list.html', context)

@is_admin_required
@login_required
def generate_radar_chart(request, id):
    questionnaire = get_object_or_404(Questionnaire, pk=id)
    matier = questionnaire.matier

    # Fetch Datainserted objects for the specific questionnaire and matier
    anne=AnneeScolaire.objects.get(current=True)
    data_objects = Datainserted.objects.filter(questionnaire=questionnaire, matier_code=matier.code).filter(anne=anne)

    sections_to_exclude = list(Section.objects.all().order_by('-order')[:4])

    # Use Python to filter out the excluded sections
    data_objects = [obj for obj in data_objects if obj.section not in sections_to_exclude]

    # Group data by section
    sections_data = {}
    for obj in data_objects:
        if obj.section not in sections_data:
            sections_data[obj.section] = {'pourcentage': 0}
        sections_data[obj.section]['pourcentage'] = Datainserted.calculate_percentage_for_section(questionnaire, obj.section)

    # Prepare data for radar chart
    data = {'group': [section.name for section in sections_data.keys()]}
    data['pourcentage'] = [sections_data[section]['pourcentage'] * 100 for section in sections_data.keys()]

    df = pd.DataFrame(data)

    # Plot the radar chart
    plt.switch_backend('Agg')  # Use Agg backend for non-GUI environment

    N = len(df['group'])
    angles = [n / float(N) * 2 * pi for n in range(N)]
    angles += angles[:1]

    ax = plt.subplot(111, polar=True)
    plt.title(f'{questionnaire.name} - {matier.code}')

    def add_line_breaks(name):
        words = name.split(' ')
        if len(words) > 4:
            words.insert(4, '\n')
        return ' '.join(words)

    # Add line breaks for every fourth space in section names
    tick_labels = [add_line_breaks(name) for name in df['group']]
    plt.xticks(angles[:-1], tick_labels, color='red', size=8)
    ax.set_rlabel_position(0)

    plt.yticks([20, 40, 60, 80, 100], ["20", "40", "60", "80", "100"], color="grey", size=7)
    plt.ylim(0, 100)

    values = df['pourcentage'].values.flatten().tolist()
    values += values[:1]
    ax.plot(angles, values, linewidth=3, linestyle='solid')
    ax.fill(angles, values, 'b', alpha=0.1)

    # Save the plot to a BytesIO buffer
    image_stream = io.BytesIO()
    plt.savefig(image_stream, format='png')
    plt.close()

    # Rewind the buffer to the beginning
    image_stream.seek(0)

    # Create a Django response with the image
    response = HttpResponse(image_stream.read(), content_type='image/png')
    response['Content-Disposition'] = f'attachment; filename={matier.code}.png'

    return response

@is_admin_required
@login_required
def view_radar_chart(request, id):
    questionnaire = get_object_or_404(Questionnaire, pk=id)
    matier = questionnaire.matier

    # Fetch Datainserted objects for the specific questionnaire and matier
    anne=AnneeScolaire.objects.get(current=True)
    data_objects = Datainserted.objects.filter(questionnaire=questionnaire, matier_code=matier.code).filter(anne=anne)
    sections_to_exclude = list(Section.objects.all().order_by('-order')[:4])

    # Use Python to filter out the excluded sections
    data_objects = [obj for obj in data_objects if obj.section not in sections_to_exclude]

    # Group data by section
    sections_data = {}
    for obj in data_objects:
        if obj.section not in sections_data:
            sections_data[obj.section] = {'pourcentage': 0}
        sections_data[obj.section]['pourcentage'] = Datainserted.calculate_percentage_for_section(questionnaire, obj.section)

    # Prepare data for radar chart
    data = {'group': [section.name for section in sections_data.keys()]}
    data['pourcentage'] = [sections_data[section]['pourcentage'] * 100 for section in sections_data.keys()]

    df = pd.DataFrame(data)

    # Plot the radar chart
    plt.switch_backend('Agg')  # Use Agg backend for non-GUI environment

    N = len(df['group'])
    angles = [n / float(N) * 2 * pi for n in range(N)]
    angles += angles[:1]

    ax = plt.subplot(111, polar=True)
    plt.title(f'{questionnaire.name} - {matier.code}')

    def add_line_breaks(name):
        words = name.split(' ')
        if len(words) > 4:
            words.insert(4, '\n')
        return ' '.join(words)

    # Add line breaks for every fourth space in section names
    tick_labels = [add_line_breaks(name) for name in df['group']]
    plt.xticks(angles[:-1], tick_labels, color='red', size=8)
    ax.set_rlabel_position(0)

    plt.yticks([20, 40, 60, 80, 100], ["20", "40", "60", "80", "100"], color="grey", size=7)
    plt.ylim(0, 100)

    values = df['pourcentage'].values.flatten().tolist()
    values += values[:1]
    ax.plot(angles, values, linewidth=3, linestyle='solid')
    ax.fill(angles, values, 'b', alpha=0.1)

    # Save the plot to a BytesIO buffer
    image_stream = io.BytesIO()
    plt.savefig(image_stream, format='png')
    plt.close()

    # Rewind the buffer to the beginning
    image_stream.seek(0)

    # Create a Django response with the image
    response = HttpResponse(image_stream.read(), content_type='image/png')

    return response

@is_admin_required
@login_required
def students_not_filled_questionnaire(request, semester_id):
    semester = get_object_or_404(Semestre, pk=semester_id)

    # Get all questionnaires associated with the semester
    questionnaires = Questionnaire.objects.filter(semester=semester)

    all_students = Etudiant.objects.filter(niveau=semester.niveau)

    # Get all matiers related to the questionnaires
    matiers_for_questionnaires = Matier.objects.filter(questionnaire__in=questionnaires).distinct()

    # Get Datainserted records for all questionnaires
    anne=AnneeScolaire.objects.get(current=True)
    data_for_questionnaires = Datainserted.objects.filter(questionnaire__in=questionnaires).filter(anne=anne)
    
    # Create a dictionary to store the filled status for each student and matier
    filled_status = {}

    for student in all_students:
        for matier in matiers_for_questionnaires:
            key = f"{student.pk}_{matier.code}"
            filled_status[key] = data_for_questionnaires.filter(etudiant=student, matier_code=matier.code).exists()

    context = {
        'questionnaires': questionnaires,
        'students_not_filled_questionnaire': all_students,
        'matiers_not_filled': matiers_for_questionnaires,
        'filled_status': filled_status,
    }

    return render(request, 'students_not_filled_questionnaire.html', context)

import xlsxwriter
@is_admin_required
@login_required
def students_not_filled_pdf(request, semester_id):
    semester = get_object_or_404(Semestre, pk=semester_id)

    # Get all questionnaires associated with the semester
    questionnaires = Questionnaire.objects.filter(semester=semester)

    all_students = Etudiant.objects.filter(niveau=semester.niveau)

    # Get all matiers related to the questionnaires
    matiers_for_questionnaires = Matier.objects.filter(questionnaire__in=questionnaires).distinct()

    # Get Datainserted records for all questionnaires
    anne=AnneeScolaire.objects.get(current=True)
    data_for_questionnaires = Datainserted.objects.filter(questionnaire__in=questionnaires).filter(anne=anne)

    # Create a dictionary to store the filled status for each student and matier
    filled_status = {}

    for student in all_students:
        for matier in matiers_for_questionnaires:
            key = f"{student.pk}_{matier.code}"
            filled_status[key] = data_for_questionnaires.filter(etudiant=student, matier_code=matier.code).exists()

    # Create an Excel workbook and add a worksheet
    output = io.BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    # Write headers
    worksheet.write('A1', 'Student')
    for idx, matier in enumerate(matiers_for_questionnaires, start=1):
        worksheet.write(0, idx, matier.name)

    # Write data
    for i, student in enumerate(all_students, start=1):
        worksheet.write(i, 0, student.Matricul)
        for j, matier in enumerate(matiers_for_questionnaires, start=1):
            key = f"{student.pk}_{matier.code}"
            if filled_status[key]:
                worksheet.write(i, j, 'OK')
            else:
                worksheet.write(i, j, 'Non')

    # Close the workbook
    workbook.close()

    # Prepare the response
    output.seek(0)
    response = HttpResponse(content=output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=students_not_filled.xlsx'

    return response


@is_admin_required
@login_required
def questionnaire_list2(request):
    questionnaires = Questionnaire.objects.all()
    context = {'questionnaires': questionnaires}
    return render(request, 'questionnaire_list2.html', context)


# section

class SectionListView(View):
    def get(self, request):
        sections = Section.objects.all()
        return render(request, 'section_list.html', {'sections': sections})

class SectionCreateView(View):
    def get(self, request):
        form = SectionForm()
        return render(request, 'section_form.html', {'form': form})

    def post(self, request):
        form = SectionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('section_list')
        return render(request, 'section_form.html', {'form': form})

class SectionUpdateView(View):
    def get(self, request, pk):
        section = get_object_or_404(Section, pk=pk)
        form = SectionForm(instance=section)
        return render(request, 'section_form.html', {'form': form, 'section': section})

    def post(self, request, pk):
        section = get_object_or_404(Section, pk=pk)
        form = SectionForm(request.POST, instance=section)
        if form.is_valid():
            form.save()
            return redirect('section_list')
        return render(request, 'section_form.html', {'form': form, 'section': section})

class SectionDeleteView(View):
    def get(self, request, pk):
        section = get_object_or_404(Section, pk=pk)
        return render(request, 'section_confirm_delete.html', {'section': section})

    def post(self, request, pk):
        section = get_object_or_404(Section, pk=pk)
        section.delete()
        return redirect('section_list')


# question

class QuestionListView(View):
      def get(self, request):
          questions = Question.objects.all()
          return render(request, 'question_list.html', {'questions': questions})

class QuestionCreateView(View):
    def get(self, request):
        form = QuestionForm()
        return render(request, 'question_form.html', {'form': form})

    def post(self, request):
        form = QuestionForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('question_list')
        return render(request, 'question_form.html', {'form': form})

class QuestionUpdateView(View):
    def get(self, request, pk):
        question = get_object_or_404(Question, pk=pk)
        form = QuestionForm(instance=question)
        return render(request, 'question_form.html', {'form': form, 'question': question})

    def post(self, request, pk):
        question = get_object_or_404(Question, pk=pk)
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            return redirect('question_list')
        return render(request, 'question_form.html', {'form': form, 'question': question})

class QuestionDeleteView(View):
    def get(self, request, pk):
        question = get_object_or_404(Question, pk=pk)
        question.delete()
        return redirect('question_list')

  # data email

    # views.py


class EmailtableListView(View):
    def get(self, request):
        emails = Emailtable.objects.all()
        return render(request, 'emailtable_list.html', {'emails': emails})

class EmailtableCreateView(View):
    def get(self, request):
        form = EmailtableForm()
        return render(request, 'emailtable_form.html', {'form': form})

    def post(self, request):
        form = EmailtableForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('emailtable_list')
        return render(request, 'emailtable_form.html', {'form': form})

class EmailtableUpdateView(View):
    def get(self, request, pk):
        email = get_object_or_404(Emailtable, pk=pk)
        form = EmailtableForm(instance=email)
        return render(request, 'emailtable_form.html', {'form': form, 'email': email})

    def post(self, request, pk):
        email = get_object_or_404(Emailtable, pk=pk)
        form = EmailtableForm(request.POST, instance=email)
        if form.is_valid():
            form.save()
            return redirect('emailtable_list')
        return render(request, 'emailtable_form.html', {'form': form, 'email': email})

class EmailtableDeleteView(View):
    def get(self, request, pk):
        email = get_object_or_404(Emailtable, pk=pk)
        return render(request, 'emailtable_confirm_delete.html', {'email': email})

    def post(self, request, pk):
        email = get_object_or_404(Emailtable, pk=pk)
        email.delete()
        return redirect('emailtable_list')

@is_admin_required
@login_required
def commentaires(request,id):
    sections_to_rest = Section.objects.all().order_by('-order')[:4]
    sections_to_rest_ids = [section.id for section in sections_to_rest]  # List of IDs
    anne = AnneeScolaire.objects.get(current=True)
    data = Datainserted.objects.exclude(matier_code__isnull=True).exclude(matier_code__exact='').filter(section__id__in=sections_to_rest_ids).order_by('section__name', 'section__order').filter(matier_code=id).filter(anne=anne).distinct()

    return render(request,'commentaires.html',{'data':data,'id':id})

from openpyxl.utils.dataframe import dataframe_to_rows
@is_admin_required
@login_required
def commentaires_download(request, id):
    sections_to_rest = Section.objects.all().order_by('-order')[:4]
    anne = get_object_or_404(AnneeScolaire, current=True)
    matier = get_object_or_404(Matier, code=id)

    # Use a list to store data as dictionaries
    data_list = []

    for section in sections_to_rest:
        data = Datainserted.objects.filter(
            matier_code=id,
            section=section,
            anne=anne
        ).order_by('question__text').distinct()

        if not data.exists():
            continue

        # Add section header
        data_list.append({
            'Content': f"--- Section: {section.name} ---",
        })

        for item in data:
            data_list.append({
                'Content': item.question.text,
            })
            data_list.append({
                'Content': item.selected_choice,
            })

    # Convert the list of dictionaries to a DataFrame
    df = pd.DataFrame(data_list)

    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # Write DataFrame to worksheet
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    # Apply bold formatting to section headers
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).startswith("---"):
                cell.font = openpyxl.styles.Font(bold=True)

    # Create Excel file in-memory
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)

    # Set up response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=comments_{matier.name}.xlsx'

    # Write the buffer to the response
    response.write(excel_buffer.getvalue())

    return response



@is_admin_required
@login_required
def matier_list(request,id):
    matier_list = Matier.objects.filter(semester=id)
    context = {'matier_list': matier_list}
    return render(request, 'matier_liste.html', context)

@is_admin_required
@login_required
def semester_list(request):
    semester_list = Semestre.objects.all()
    context = {'semester_list': semester_list}
    return render(request, 'semester_list.html', context)


@is_admin_required
@login_required
def delete_comment(request, comment_id , id):
    comment = get_object_or_404(Datainserted, pk=comment_id)
    comment.delete()
    return redirect('commentaires',id)

@is_admin_required
@login_required
def semester_list2(request):
    semester_list = Semestre.objects.all()
    context = {'semester_list': semester_list}
    return render(request, 'semester_graph.html', context)


@is_admin_required
@login_required
def semester_list3(request):
    semester_list = Semestre.objects.all()
    context = {'semester_list': semester_list}
    return render(request, 'semester_notfilled.html', context)


@is_admin_required
@login_required
def semester_list4(request):
    semester_list = Semestre.objects.all()
    context = {'semester_list': semester_list}
    return render(request, 'semester_notfilled_mat.html', context)

import textwrap

@is_admin_required
@login_required
def generate_bar_chart(request, semester_id):
    semester = get_object_or_404(Semestre, pk=semester_id)
    # Fetch all questionnaires for the specific semester
    questionnaires = Questionnaire.objects.filter(semester=semester)

    # Fetch Datainserted objects for all questionnaires in the semester
    anne = AnneeScolaire.objects.get(current=True)
    data_objects = Datainserted.objects.filter(questionnaire__in=questionnaires).filter(anne=anne)


    sections_to_exclude = list(Section.objects.all().order_by('-order')[:4])

    # Use Python to filter out the excluded sections
    data_objects = [obj for obj in data_objects if obj.section not in sections_to_exclude]

    # Group data by matier
    matiers_data = {}
    for obj in data_objects:
        matier_code = obj.matier_code
        matier = get_object_or_404(Matier, code=matier_code)

        if matier not in matiers_data:
            matiers_data[matier] = {'pourcentage': 0}

        matiers_data[matier]['pourcentage'] = Datainserted.calculate_percentage_for_section(obj.questionnaire, obj.section)

    # Prepare data for bar chart
    if not matiers_data:
        # If there is no data, return an appropriate response
        return HttpResponse("No data available for this semester.")

    # Convert the dictionary to a list of tuples and sort it by percentage
    sorted_data = sorted(matiers_data.items(), key=lambda item: item[1]['pourcentage'])

    # Prepare the sorted data for the DataFrame
    data = {
        'matier': [matier.code for matier, _ in sorted_data],
        'pourcentage': [pourcentage['pourcentage'] * 100 for _, pourcentage in sorted_data]
    }

    # Add line breaks to long matier names
    max_width = 20  # You can adjust the maximum width as needed
    data['matier'] = ['\n'.join(textwrap.wrap(name, max_width)) for name in data['matier']]

    df = pd.DataFrame(data)

    # Use seaborn style for a more sophisticated look
    sns.set(style="whitegrid")

    # Create a figure with increased size and resolution
    plt.figure(figsize=(10, 6), dpi=100)

    # Plot the bar chart with a nicer color palette
    ax = sns.barplot(x='matier', y='pourcentage', data=df, color='#276bc9')

    # Add a red line at 60%
    plt.axhline(y=60, color='red', linewidth=1.5, linestyle='--')

    # Annotate bars with their respective percentages
    for p in ax.patches:
        ax.annotate(format(p.get_height(), '.1f') + '%',
                    (p.get_x() + p.get_width() / 2., p.get_height()),
                    ha='center', va='center',
                    xytext=(0, 9),
                    textcoords='offset points')

    plt.title(f'Qualité - {semester.nom_semestre}', fontsize=16)
    plt.xlabel('Élément du module', fontsize=14)
    plt.ylabel('Pourcentage (%)', fontsize=14)
    plt.xticks(rotation=45, ha='right', fontsize=12)
    plt.yticks(fontsize=12)
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))  # Ensure y-axis ticks are integer

    # Save the plot to a BytesIO buffer
    image_stream = io.BytesIO()
    plt.savefig(image_stream, format='png', bbox_inches='tight')
    plt.close()

    # Rewind the buffer to the beginning
    image_stream.seek(0)

    # Create a Django response with the image and set it for download
    response = HttpResponse(image_stream.read(), content_type='image/png')
    response['Content-Disposition'] = f'attachment; filename="Qualite_{semester.nom_semestre}.png"'

    return response


import textwrap

@is_admin_required
@login_required
def view_bar_chart(request, id):
    semester = get_object_or_404(Semestre, pk=id)
    # Fetch all questionnaires for the specific semester
    questionnaires = Questionnaire.objects.filter(semester=semester)

    # Fetch Datainserted objects for all questionnaires in the semester
    anne = AnneeScolaire.objects.get(current=True)
    data_objects = Datainserted.objects.filter(questionnaire__in=questionnaires).filter(anne=anne)


    sections_to_exclude = list(Section.objects.all().order_by('-order')[:4])

    # Use Python to filter out the excluded sections
    data_objects = [obj for obj in data_objects if obj.section not in sections_to_exclude]

    # Group data by matier
    matiers_data = {}
    for obj in data_objects:
        matier_code = obj.matier_code
        matier = get_object_or_404(Matier, code=matier_code)

        if matier not in matiers_data:
            matiers_data[matier] = {'pourcentage': 0}

        matiers_data[matier]['pourcentage'] = Datainserted.calculate_percentage_for_section(obj.questionnaire, obj.section)

    # Prepare data for bar chart
    if not matiers_data:
        # If there is no data, return an appropriate response
        return HttpResponse("No data available for this semester.")

    # Convert the dictionary to a list of tuples and sort it by percentage
    sorted_data = sorted(matiers_data.items(), key=lambda item: item[1]['pourcentage'])

    # Prepare the sorted data for the DataFrame
    data = {
        'matier': [matier.code for matier, _ in sorted_data],
        'pourcentage': [pourcentage['pourcentage'] * 100 for _, pourcentage in sorted_data]
    }

    # Add line breaks to long matier names
    max_width = 20  # You can adjust the maximum width as needed
    data['matier'] = ['\n'.join(textwrap.wrap(name, max_width)) for name in data['matier']]

    df = pd.DataFrame(data)

    # Use seaborn style for a more sophisticated look
    sns.set(style="whitegrid")

    # Create a figure with increased size and resolution
    plt.figure(figsize=(10, 6), dpi=100)

    # Plot the bar chart with a nicer color palette
    color = sns.color_palette("viridis", as_cmap=False)[1]
    ax = sns.barplot(x='matier', y='pourcentage', data=df, color='#276bc9')

    # Add a red line at 60%
    plt.axhline(y=60, color='red', linewidth=1.5, linestyle='--')

    # Annotate bars with their respective percentages
    for p in ax.patches:
        ax.annotate(format(p.get_height(), '.1f') + '%',
                    (p.get_x() + p.get_width() / 2., p.get_height()),
                    ha='center', va='center',
                    xytext=(0, 9),
                    textcoords='offset points')

    plt.title(f'Qualité - {semester.nom_semestre}', fontsize=16)
    plt.xlabel('Élément du module', fontsize=14)
    plt.ylabel('Pourcentage (%)', fontsize=14)
    plt.xticks(rotation=45, ha='right', fontsize=12)
    plt.yticks(fontsize=12)
    ax.yaxis.set_major_locator(MaxNLocator(integer=True))  # Ensure y-axis ticks are integer

    # Save the plot to a BytesIO buffer
    image_stream = io.BytesIO()
    plt.savefig(image_stream, format='png', bbox_inches='tight')
    plt.close()

    # Rewind the buffer to the beginning
    image_stream.seek(0)

    # Create a Django response with the image
    response = HttpResponse(image_stream.read(), content_type='image/png')

    return response




class DownloadAllRadarChartsView(View):
    def get(self, request, semester_id):
        # Fetch all questionnaires for the specific semester
        questionnaires = Questionnaire.objects.filter(semester_id=semester_id)
        semesters = Semestre.objects.get(pk=semester_id)
        semester = semesters.nom_semestre
        # Create a BytesIO buffer to store the zip file
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
            # Generate radar chart for each questionnaire
            for questionnaire in questionnaires:
                radar_chart = self.generate_radar_chart(questionnaire)
                # Save the chart to a BytesIO buffer
                chart_buffer = io.BytesIO()
                radar_chart.savefig(chart_buffer, format='png')
                chart_buffer.seek(0)
                # Add the chart to the zip file with a unique name
                zip_file.writestr(f"{questionnaire.matier.code}_radar_chart.png", chart_buffer.read())

        # Rewind the zip buffer to the beginning
        zip_buffer.seek(0)

        # Create a Django response with the zip file
        response = HttpResponse(zip_buffer.read(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename=radar_charts-{semester}.zip'

        return response

    def generate_radar_chart(self, questionnaire):
        matier = questionnaire.matier

        # Fetch Datainserted objects for the specific questionnaire and matier
        anne=AnneeScolaire.objects.get(current=True)
        data_objects = Datainserted.objects.filter(questionnaire=questionnaire, matier_code=matier.code).filter(anne=anne)

        sections_to_exclude = list(Section.objects.all().order_by('-order')[:4])

        # Use Python to filter out the excluded sections
        data_objects = [obj for obj in data_objects if obj.section not in sections_to_exclude]

        # Group data by section
        sections_data = {}
        for obj in data_objects:
            if obj.section not in sections_data:
                sections_data[obj.section] = {'pourcentage': 0}
            sections_data[obj.section]['pourcentage'] = Datainserted.calculate_percentage_for_section(questionnaire, obj.section)

        # Prepare data for radar chart
        data = {'group': [section.name for section in sections_data.keys()]}
        data['pourcentage'] = [sections_data[section]['pourcentage'] * 100 for section in sections_data.keys()]

        df = pd.DataFrame(data)

        # Plot the radar chart
        plt.switch_backend('Agg')  # Use Agg backend for non-GUI environment

        N = len(df['group'])
        angles = [n / float(N) * 2 * pi for n in range(N)]
        angles += angles[:1]

        ax = plt.subplot(111, polar=True)
        plt.title(f'{questionnaire.name} - {matier.code}')

        def add_line_breaks(name):
            words = name.split(' ')
            if len(words) > 4:
                words.insert(4, '\n')
            return ' '.join(words)

        # Add line breaks for every fourth space in section names
        tick_labels = [add_line_breaks(name) for name in df['group']]
        plt.xticks(angles[:-1], tick_labels, color='red', size=8)
        ax.set_rlabel_position(0)

        plt.yticks([20, 40, 60, 80, 100], ["20", "40", "60", "80", "100"], color="grey", size=7)
        plt.ylim(0, 100)

        values = df['pourcentage'].values.flatten().tolist()
        values += values[:1]
        ax.plot(angles, values, linewidth=3, linestyle='solid')
        ax.fill(angles, values, 'b', alpha=0.1)

        return plt


@is_admin_required
@login_required
def data_inserted_table(request):
    anne=AnneeScolaire.objects.get(current=True)
    data_objects = Datainserted.objects.all().filter(anne=anne)
    context = {'data_objects': data_objects}
    return render(request, 'data_inserted_table.html', context)


@is_admin_required
@login_required
def import_matiere(request):
    if request.method == 'POST':
        form = ImportMatiereForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            df = pd.read_excel(excel_file, engine='openpyxl')

            for index, row in df.iterrows():
                semestre_name = row['nom_semestre']

                # Try to get or create the Semestre instance based on the provided semestre name
                semestre_instance= Semestre.objects.get(
                    nom_semestre=semestre_name
                )

                Matier.objects.create(
                    code=row['code'],
                    name=row['name'],
                    semester=semestre_instance
                )

            return redirect('matier_table')  # Redirect to a success page after importing
    else:
        form = ImportMatiereForm()

    return render(request, 'import_matiere.html', {'form': form})